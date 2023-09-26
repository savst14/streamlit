import streamlit as st
import os
import pandas as pd
from pyacadcom import AutoCAD, AcadPoint
from openpyxl import load_workbook
from collections import Counter
import pythoncom

# pythoncom.CoInitializeEx(0)

def get_info_exel(excel_file):
    # excel_file = file_exel.replace('/', '\\')
    sheets = load_workbook(excel_file, read_only=True).sheetnames
    # print(sheets)
    panel_counter = 1  # начинаем считать панели
    panels_list = []  # list of panels
    for sheet in range(1, len(sheets)):  # don't take a first list coz it's technical list
        if sheets[sheet] == 'Example':  # don't take a Example list
            continue
        # print('sh-', sheet)
        panel_list = []  # ИНФО О ПАНЕЛИ: лист c кол-вом фидеров и список с фидерами
        df = pd.read_excel(excel_file,
                           sheet_name=sheets[sheet])
        # feeder counter
        feeders_nums = set()
        for feeder_num in df['Группа']:
            try:
                feeders_nums.add(int(feeder_num))
            except:
                continue
        quantity = len(feeders_nums)
        # print(feeders_nums)
        # take feeder info
        feeders_list = []  # list of feeder info of panel
        sum_power = 0  # установленная мощность float Вт
        i_cb_max_set = set()  # design power (расчетная мощность)
        module_counter = 0
        place_counter = quantity

        # get information about feeders
        for group in range(1, quantity):  # было range(1, quantity+1)
            feeder_dict = {}
            feeder_name = f'Гр.{group}'
            feeder_dict.setdefault('feeder', feeder_name)
            # U,В - обязательное значение
            print(feeder_name, 'U,В')
            U_B = get_column_exel(df, group, 'U,В')[0]  # for cable and pole of CB and current
            try:
                U = int(U_B)
            except:  # вписать в отчет об ошибке
                feeder_dict.setdefault('pnom', '')
                feeder_dict.setdefault('Kc', '')
                feeder_dict.setdefault('cos', '')
                feeder_dict.setdefault('Iном, А', '')
                feeder_dict.setdefault('load_name', '')
                feeder_dict.setdefault('cb', 'QF')  # нужно указать, чтобы вышел пустой фидер с типом QF
                feeder_dict.setdefault('i_cb_set', '')
                feeder_dict.setdefault('cable', '')
                feeder_dict.setdefault('len', '')
                feeder_dict.setdefault('lay', '')
                feeder_dict.setdefault('lay_l', '')
                feeder_dict.setdefault('v_loss', '')
                feeder_dict.setdefault('sf_counter', 0)  # нужно указать, чтобы вышел пустой фидер без подфидеров
                continue  # скрипт переходит к новой группе
            # PNOM - subfeeder - start
            print(feeder_name, 'PNOM - subfeeder')
            pnom_subfeeder = []  # лист с мощностью каждого подфидера в формате float в Вт
            dicts_list = [{}, {}, {}]
            sf_counter = 0
            for i in range(1, 4):
                try:
                    power_1_2_3 = get_float_exel(df, group, f'{i}_Руст, Вт')  # вернул float
                    dicts_list[i - 1].setdefault('pnom', format(power_1_2_3 / 1000, '.2f'))
                    pnom_subfeeder.append(power_1_2_3)
                    sf_counter += 1
                except:
                    pnom_subfeeder.append(0)
                    dicts_list[i - 1].setdefault('pnom', 0)
            if sf_counter > 0:
                place_counter += sf_counter - 1
            feeder_dict.setdefault('sf_counter', sf_counter)  # указали количество подфидеров в списке фидера
            # U,В - subfeeder
            print(feeder_name, 'U,В - subfeeder', pnom_subfeeder)
            u_subfeeder = []
            for i in range(1, 4):
                if pnom_subfeeder[i - 1] == 0:  # подфидера нет
                    u_subfeeder.append(0)
                else:  # если фидер есть, то находим напряжение
                    try:
                        u_1_2_3 = get_float_exel(df, group, f'{i}_U,В')
                        u_subfeeder.append(int(u_1_2_3))
                    except:
                        u_subfeeder.append(U)
            # PNOM
            print(feeder_name, 'PNOM', pnom_subfeeder)
            try:
                feeder_power = get_float_exel(df, group, 'Руст, Вт')  # float Вт
                pnom = format(feeder_power / 1000, '.2f')  # в кВт
            except:
                feeder_power = sum(pnom_subfeeder)  # сумма мощностей подфидеров
                if feeder_power == 0:  # сумма списка
                    pnom = ''  # если значения нет, значит это резерв
                else:
                    pnom = format(feeder_power / 1000, '.2f')
            sum_power += feeder_power
            feeder_dict.setdefault('pnom', pnom)
            # Kc
            if pnom == '':  # если это резерв, то нам не нужны надписи на фидоре
                feeder_use = ''
            else:
                try:
                    feeder_use = get_float_exel(df, group, 'Kc')
                    # des_power += feeder_power * feeder_use
                except:
                    feeder_use = 1
                    # des_power += feeder_power
            feeder_dict.setdefault('Kc', feeder_use)  # нам нужно это значение для отчета
            # COS
            if pnom == '':  # если это резерв, то нам не нужны надписи на фидоре
                feeder_cos = ''
                cos_subfeeder = [0, 0, 0]  # не удалять
            else:
                try:
                    feeder_cos = get_float_exel(df, group, 'Cos')
                    # feeder_cos = format(feeder_cos, '.2f')
                except:
                    feeder_cos = 0.80
                # feeder_dict.setdefault('cos', feeder_cos)
            # COS - subfeeder
            print(feeder_name, 'COS - subfeeder')
            cos_subfeeder = []
            for i in range(1, 4):
                if pnom_subfeeder[i - 1] == 0:  # подфидера нет
                    cos_1_2_3 = 0
                else:  # если фидер есть, то находим
                    try:
                        cos_1_2_3 = get_float_exel(df, group, f'{i}_Cos')
                    except:
                        cos_1_2_3 = 0.8
                cos_subfeeder.append(cos_1_2_3)
            # Iрасч
            print(feeder_name, 'Iрасч')
            if pnom == '':  # значит что резерв
                feeder_current = ''
                current_subfeeder = [0, 0, 0]
                feeder_dict.setdefault('Iрасч, А', feeder_current)
            else:
                # Iрасч - subfeeder - start
                print(feeder_name, 'Iрасч - subfeeder')
                current_subfeeder = []
                for i in range(1, 4):
                    if pnom_subfeeder[i - 1] == 0:  # подфидера нет
                        current_subfeeder.append(0)
                    else:  # если фидер есть, то находим
                        try:
                            current_1_2_3 = get_float_exel(df, group, f'{i}_Iрасч, А')
                        except:  # если фидер есть, то находим
                            # print(u_subfeeder[i - 1])
                            if u_subfeeder[i - 1] <= 230:
                                current_1_2_3 = \
                                    pnom_subfeeder[i - 1] / cos_subfeeder[i - 1] / u_subfeeder[i - 1]  # float
                            else:
                                current_1_2_3 = \
                                    pnom_subfeeder[i - 1] / cos_subfeeder[i - 1] / 1.73205081 / u_subfeeder[i - 1]
                        current_subfeeder.append(current_1_2_3)
                        dicts_list[i - 1].setdefault('Iрасч, А', format(current_1_2_3, '.1f'))
                # Iрасч - subfeeder - end
                print(feeder_name, 'Iрасч - subfeeder - end')
                try:
                    feeder_current = get_float_exel(df, group, 'Iрасч, А')
                except:  # значит будем вычислять
                    if U <= 230:
                        feeder_current = feeder_power * feeder_use / feeder_cos / U
                    else:
                        feeder_current = feeder_power * feeder_use / feeder_cos / 1.73205081 / U
                feeder_dict.setdefault('Iрасч, А', format(feeder_current, '.1f'))
            # CB
            print(feeder_name, 'CB')
            try:
                type_cb = get_column_exel(df, group, 'Автомат')[0]  # QF, QFD, QS and add QFS (независ. расцепитель)
            except:
                type_cb = 'QF'
            group_c = group
            if U <= 230:  # значение напряжения введено, значит нужно что то поставить
                if type_cb == 'QF':
                    pole = 1
                    module_counter += 1
                elif type_cb == 'QFD':
                    pole = 2
                    module_counter += 2
                elif type_cb == 'QS':
                    pole = 2  # примем что используем двухполюсные ВН
                    module_counter += 2
                elif type_cb == 'РН-QF':
                    pole = 1  # автомат остается однополюсным
                    module_counter += 2  # но к автомату ставим РН
                else:
                    type_cb = 'QF'
                    pole = 1
                    module_counter += 1
            else:
                if type_cb == 'QF':
                    pole = 3
                    module_counter += 3
                elif type_cb == 'QFD' or type_cb == 'QS':
                    pole = 4  # примем что используем четырехполюсные ВН
                    module_counter += 4
                elif type_cb == 'РН-QF':
                    pole = 3  # автомат остается трехполюсным
                    module_counter += 4  # но к автомату ставим РН
                else:
                    type_cb = 'QF'
                    pole = 3
                    module_counter += 3
            feeder_dict.setdefault('type_cb', type_cb)  # нужно для выбора блока фидера
            feeder_dict.setdefault('cb', f'{type_cb}{group_c}({pole}p)')
            # I_SB_SET
            try:
                i_cb_set = int(get_float_exel(df, group, 'Уставка'))
                i_cb_max_set.add(i_cb_set)  # добавляем в список автоматов щита, чтобы подобрать вводной потом
            except:  # вычисляем
                if pnom == '':  # значит что резерв
                    i_cb_set = ''
                else:
                    i_cb_set = calc_cb(feeder_current)
                    i_cb_max_set.add(i_cb_set)
            feeder_dict.setdefault('i_cb_set', i_cb_set)
            # N_PLAN - subfeeder
            print(feeder_name, 'N_PLAN - subfeeder')
            for i in range(1, 4):
                if pnom_subfeeder[i - 1] == 0:  # подфидера нет
                    continue
                else:  # если фидер есть, то находим
                    try:
                        n_plan_1_2_3 = get_column_exel(df, group, f'{i}_N по плану')[0]
                    except:  # если фидер есть, то находим
                        n_plan_1_2_3 = ''
                    if type(n_plan_1_2_3) == str:
                        dicts_list[i - 1].setdefault('n_plan', n_plan_1_2_3)
                    else:
                        dicts_list[i - 1].setdefault('n_plan', '')
            # N_PLAN
            try:
                n_plan = get_column_exel(df, group, 'N по плану')[0]
            except:
                n_plan = ''
            if type(n_plan) == str:
                feeder_dict.setdefault('n_plan', n_plan)
            else:
                feeder_dict.setdefault('n_plan', '')
            # LOAD_NAME - subfeeder
            print(feeder_name, 'LOAD_NAME - subfeeder')
            for i in range(1, 4):
                if pnom_subfeeder[i - 1] == 0:  # подфидера нет
                    continue
                else:  # если фидер есть, то находим
                    try:
                        load_name_1_2_3 = get_column_exel(df, group, f'{i}_Нагрузка')[0]
                    except:  # если фидер есть, то находим
                        load_name_1_2_3 = ''
                    if type(load_name_1_2_3) == str:
                        dicts_list[i - 1].setdefault('load_name', load_name_1_2_3)
                    else:
                        dicts_list[i - 1].setdefault('load_name', '')
            # LOAD_NAME
            try:
                load_name = get_column_exel(df, group, 'Нагрузка')[0]
            except:
                if pnom == '':
                    load_name = 'Резерв'
                else:
                    load_name = ''
            if type(load_name) == str:
                feeder_dict.setdefault('load_name', load_name)
            else:
                feeder_dict.setdefault('load_name', '')
            # LEN - subffeder
            print(feeder_name, 'LEN - subffeder')
            length_subfeeder = []  # для длин кабелей фидеров (если нет, то сечение 0)
            for i in range(1, 4):
                if pnom_subfeeder[i - 1] == 0:  # подфидера нет
                    length_subfeeder.append(0)
                    continue
                else:
                    try:  # взяли значение
                        length_1_2_3 = get_column_exel(df, group, f'{i}_Длина, м')[0]
                    except:  # столбец отсутсвует - вводим пустоту
                        length_1_2_3 = ''
                    try:  # если это число, то записали
                        length_subfeeder.append(int(length_1_2_3))
                        dicts_list[(i - 1)].setdefault('len', int(length_1_2_3))
                    except:  # значит это компл. или (см.проект...) или пустота
                        if type(length_1_2_3) == str:  # компл. или учтено в проекте
                            length_subfeeder.append(0)
                            dicts_list[(i - 1)].setdefault('len', length_1_2_3)
                        else:  # значение отсутсвует - вводим пустоту
                            length_subfeeder.append(0)
                            dicts_list[(i - 1)].setdefault('len', '')
            # LEN
            print(feeder_name, 'LEN')
            try:
                length = get_column_exel(df, group, 'Длина, м')[0]  # проверяем, есть ли какое-то значение
            except:
                length = ''  # если нет даже столбца
            try:
                1 / int(length)  # выдаст ошибку и уйдет в исключения если это не число и не ноль
                feeder_dict.setdefault('len', int(length))  # записали
            except:
                if type(length) == str:  # написали "комлектно", см. проект
                    feeder_dict.setdefault('len', length)
                else:  # длину не указали
                    feeder_dict.setdefault('len', '')
                feeder_dict.setdefault('cable', '')
                feeder_dict.setdefault('v_loss', '')
                feeder_dict.setdefault('lay', '')
                feeder_dict.setdefault('lay_l', '')
                feeder_dict.setdefault('sf_counter', sf_counter)
                for i in range(1, 4):
                    if pnom_subfeeder[i - 1] == 0:  # подфидера нет
                        continue
                    else:  # добавляем подфидеры в словарь фидера
                        dicts_list[(i - 1)].setdefault('cable', '')
                        dicts_list[(i - 1)].setdefault('v_loss', '')
                        dicts_list[(i - 1)].setdefault('lay', '')
                        dicts_list[(i - 1)].setdefault('lay_l', 0)
                        feeder_dict.setdefault(f'subfeeder_{i}', dicts_list[i - 1])  # add подфидер в словарь фидера
                feeders_list.append(feeder_dict)
                continue  # переход к следующему фидеру
            # т.к могут ввести длину кабеля при этом не указать нагрузку, то
            if feeder_power == '' or feeder_power == 0:  # если нагрузки нет, то остальные поля заполняем пустыми
                feeder_dict.setdefault('cable', '')
                feeder_dict.setdefault('v_loss', '')
                feeder_dict.setdefault('lay', '')
                feeder_dict.setdefault('lay_l', '')
                feeder_dict.setdefault('sf_counter', sf_counter)
                for i in range(1, 4):
                    if pnom_subfeeder[i - 1] == 0:  # подфидера нет
                        continue
                    else:  # добавляем подфидеры в словарь фидера
                        dicts_list[(i - 1)].setdefault('cable', '')
                        dicts_list[(i - 1)].setdefault('v_loss', '')
                        dicts_list[(i - 1)].setdefault('lay', '')
                        dicts_list[(i - 1)].setdefault('lay_l', 0)
                        feeder_dict.setdefault(f'subfeeder_{i}', dicts_list[i - 1])  # add подфидер в словарь фидера
                feeders_list.append(feeder_dict)
                continue  # переход к следующему фидеру
            # LEN for dU
            try:
                length_du = get_float_exel(df, group, 'Длина для dU, м')
            except:
                length_du = length
            # CABLE_TYPE
            print(feeder_name, 'CABLE_TYPE')
            try:
                cable_t = get_column_exel(df, group, 'Кабель')[0]
            except:
                cable_t = 'ВВГнг(A)-LS'  # по умолчанию будет данный тип кабеля
            if type(cable_t) == str:  # это всегда текст. nan - это float, поэтому фильтр не пройдет
                cable_type = cable_t
            else:
                cable_type = 'ВВГнг(A)-LS'  # по умолчанию будет данный тип кабеля
            # CABLE_SIZE - subfeeder - предварительный расчет
            print(feeder_name, 'CABLE_SIZE - subfeeder')
            symbol_subfeeder = []  # предварительные сечения кабелей фидеров [CB, '=', 1] 0 - если нет фидера
            s_subfeeder = []  # предварительные сечения кабелей фидеров [2,5, 4, 2.5] 0 - если нет фидера
            d_v_subfeeder = []  # предварительные падения напряжений на фидерах [0,8, 0,8, 0.8] 0 - если нет фидера
            for i in range(1, 4):
                if length_subfeeder[i - 1] == 0:  # если фидера нет или компл., то не нужно ничего считывать и считать
                    symbol_subfeeder.append(0)
                    s_subfeeder.append(0)
                    d_v_subfeeder.append(0)
                    continue
                else:
                    try:  # проверка наличия столбца
                        cable_s_1_2_3 = get_column_exel(df, group, f'{i}_Сечение')[0]
                    except:  # иначе будем находить сечение ниже
                        cable_s_1_2_3 = None
                    try:  # если это число, то считаем сечение и падение напряжения
                        int(cable_s_1_2_3)
                        symbol_subfeeder.append(1)  # добавляем символ принудительного значения сечения
                        if cable_s_1_2_3 < 3:  # берем кабели сечением до 2,5мм2, чтобы отобразить их правильно
                            size_dv_subfeeder = [format(cable_s_1_2_3, '.1f'), calc_dv(  # считаем потери
                                u_subfeeder[i - 1], length_subfeeder[i - 1], pnom_subfeeder[i - 1],
                                1, cable_type, cable_s_1_2_3
                            )]
                        else:  # берем кабели сечением с 4мм2, чтобы отобразить их правильно
                            size_dv_subfeeder = [format(cable_s_1_2_3, '.0f'), calc_dv(  # считаем потери
                                u_subfeeder[i - 1], length_subfeeder[i - 1], pnom_subfeeder[i - 1],
                                1, cable_type, cable_s_1_2_3
                            )]
                    except:  # значение либо '=' либо не введено (столбец может быть удален)
                        if cable_s_1_2_3 == 'CB' or cable_s_1_2_3 == 'СВ':
                            symbol_subfeeder.append('CB')
                            i_cb_set_subfeeder = i_cb_set  # берем автомат с фидера
                            size_dv_subfeeder = calc_s(  # cable of subfeeder
                                u_subfeeder[i - 1], length_subfeeder[i - 1], pnom_subfeeder[i - 1],
                                1, cable_type, i_cb_set_subfeeder, 0
                            )
                        else:  # считаем сечение только по нагрузке
                            if cable_s_1_2_3 == '=':  # сечение подфидера равно значению кабеля фидера
                                symbol_subfeeder.append('=')
                            else:
                                symbol_subfeeder.append(1)
                            i_cb_set_subfeeder = calc_cb(current_subfeeder[i - 1])  # cb of subfeeder
                            size_dv_subfeeder = calc_s(  # cable of subfeeder
                                u_subfeeder[i - 1], length_subfeeder[i - 1], pnom_subfeeder[i - 1],
                                1, cable_type, i_cb_set_subfeeder, 0
                            )
                    s_subfeeder.append(size_dv_subfeeder[0])
                    d_v_subfeeder.append(size_dv_subfeeder[1])
            # CABLE_SIZE
            print(feeder_name, 'CABLE_SIZE')
            try:
                cable_s = get_float_exel(df, group, 'Сечение')
                feeder_loss = 0  # предварительное значение
            except:  # иначе подбираем кабель с учетом максимального падения напряжения на подфидерах
                cable_s_dv = calc_s(
                    U, length_du, feeder_power, feeder_use, cable_type, i_cb_set, max(d_v_subfeeder))
                cable_s = cable_s_dv[0]
                feeder_loss = cable_s_dv[1]
            if cable_s < 3:  # берем кабели сечением до 2,5мм2, чтобы отобразить их правильно
                cable_size = format(cable_s, '.1f')  # CABLE_SIZE
            else:  # берем кабели сечением с 4мм2, чтобы отобразить их правильно
                cable_size = format(cable_s, '.0f')  # CABLE_SIZE
            if U <= 230:
                cable = f'{cable_type} 3x{cable_size}'
            else:
                cable = f'{cable_type} 5x{cable_size}'
            feeder_dict.setdefault('cable', cable)
            # V_LOSS
            print(feeder_name, 'V_LOSS')
            try:
                feeder_loss = get_float_exel(df, group, 'dU')
                feeder_dict.setdefault('v_loss', format(feeder_loss, '.2f'))
            except:
                if feeder_loss > 0:
                    feeder_dict.setdefault('v_loss', format(feeder_loss, '.2f'))
                else:
                    feeder_loss = calc_dv(  # считаем потери
                        U, length_du, feeder_power, feeder_use, cable_type, cable_s)
                    feeder_dict.setdefault('v_loss', format(feeder_loss, '.2f'))
            # CABLE_SIZE и V_LOSS - subfeeder - окончательный расчет
            print(feeder_name, 'CABLE_SIZE и V_LOSS- subfeeder')
            cable_subfeeder_list = []  # лист для
            for i in range(1, 4):
                if symbol_subfeeder[i - 1] == 0:  # здесь может быть компл. поэтому заполняем до конца
                    cable_subfeeder_list.append(0)
                    dicts_list[(i - 1)].setdefault('cable', '')
                    dicts_list[(i - 1)].setdefault('v_loss', '')
                    dicts_list[(i - 1)].setdefault('lay', '')
                    dicts_list[(i - 1)].setdefault('lay_l', 0)
                elif symbol_subfeeder[i - 1] == '=':  # берем сечение фидера cable_size
                    if u_subfeeder[i - 1] <= 230:
                        cable_subfeeder = f'{cable_type} 3x{cable_size}'
                    else:
                        cable_subfeeder = f'{cable_type} 5x{cable_size}'
                    # s_subfeeder[i - 1] = cable_size  # обновил значение сечения в списке
                    cable_subfeeder_list.append(cable_subfeeder)
                    dicts_list[(i - 1)].setdefault('cable', cable_subfeeder)
                    try:
                        subfeeder_loss = get_float_exel(df, group, f'{i}_dU')
                    except:  # возможно изменение кабеля, нельзя брать из списка du для подфидеров
                        subfeeder_loss = calc_dv(
                            u_subfeeder[i - 1], length_subfeeder[i - 1], pnom_subfeeder[i - 1],
                            1, cable_type, float(cable_size)) + feeder_loss
                    dicts_list[(i - 1)].setdefault('v_loss', format(subfeeder_loss, '.2f'))
                else:  # в остальных случаях сечение уже подобрано правильно, остается только записать его и потери
                    if u_subfeeder[i - 1] <= 230:
                        cable_subfeeder = f'{cable_type} 3x{s_subfeeder[i - 1]}'
                    else:
                        cable_subfeeder = f'{cable_type} 5x{s_subfeeder[i - 1]}'
                    dicts_list[(i - 1)].setdefault('cable', cable_subfeeder)
                    cable_subfeeder_list.append(cable_subfeeder)
                    try:
                        subfeeder_loss = get_column_exel(df, group, f'{i}_dU')[0]
                        int(subfeeder_loss)
                    except:
                        subfeeder_loss = d_v_subfeeder[i - 1] + feeder_loss
                    dicts_list[(i - 1)].setdefault('v_loss', format(subfeeder_loss, '.2f'))
            # LAY
            print(feeder_name, 'LAY')
            try:
                type_lay = get_column_exel(df, group, 'Тип прокладки')[0]
            except:
                type_lay = 0  # если столбец удален, то прокладка без трубы
            if type(type_lay) == str:  # если значение введено, то это прокладка в трубе (введено значение П)
                lay = f'{type_lay}{find_lay(cable)}'  # Получается П20 или П25 и тд.
            else:
                lay = ''
            feeder_dict.setdefault('lay', lay)
            # LAY - subfeeder
            # print(feeder_name, 'LAY - subfeeder')
            for i in range(1, 4):
                if symbol_subfeeder[i - 1] == 0:
                    dicts_list[(i - 1)].setdefault('lay', '')
                    continue
                elif type(type_lay) == str:  # если значение введено, то это прокладка в трубе (введено значение П)
                    # print(f'lay_feeder_{i}', s_subfeeder[i - 1], type_lay)
                    lay_subf = f'{type_lay}{find_lay(cable_subfeeder_list[i - 1])}'  # Получается П20 или П25 и тд.
                else:
                    lay_subf = ''
                dicts_list[(i - 1)].setdefault('lay', lay_subf)
            # LAY_LENGTH
            # print(feeder_name, 'LAY_LENGTH')
            if lay == '':
                feeder_dict.setdefault('lay_l', 0)  # не будет показано на однолинейке, но будет учитываться в BOM
            else:
                feeder_dict.setdefault('lay_l', int(length - 1))
            # LAY_LENGTH - subfeeder
            for i in range(1, 4):
                if length_subfeeder[i - 1] == 0:
                    dicts_list[(i - 1)].setdefault('lay_l', 0)  # не будет показано на однолинейке
                    continue
                else:
                    dicts_list[(i - 1)].setdefault('lay_l', length_subfeeder[i - 1] - 1)
            # ADD info of subfeeders to feeder_dict
            for i in range(1, 4):
                if pnom_subfeeder[i - 1] == 0:  # подфидера нет
                    continue
                else:  # добавляем подфидеры в словарь фидера
                    feeder_dict.setdefault(f'subfeeder_{i}', dicts_list[i - 1])
            feeders_list.append(feeder_dict)
        # cчитываем вводной фидер
        feeder_dict = {}
        group = 0
        feeder_dict.setdefault('feeder', 'Ввод')
        # U_В - ввод
        # print(sheet, 'U_В - ввод')
        try:
            U = get_float_exel(df, group, 'U,В')
        except:
            U = 380
        #  PNOM - ввод
        print(sheet, 'PNOM - ввод')
        try:
            feeder_power = get_float_exel(df, group, 'Руст, Вт')
            pnom = format(feeder_power / 1000, '.2f')  # в кВт
        except:
            feeder_power = sum_power
            pnom = format(feeder_power / 1000, '.2f')
        feeder_dict.setdefault('pnom', pnom)
        # Kc - ввод
        try:
            feeder_use = get_float_exel(df, group, 'Kc')
        except:
            feeder_use = 1
        feeder_dict.setdefault('Kc', format(feeder_use, '.2f'))
        # COS-ввод
        try:
            print('ВВОД!!!!!')
            feeder_cos = get_float_exel(df, group, 'Cos')
            print('ВВОД ПРОЙДЕН!!!!')
        except:
            feeder_cos = 0.80  # тут можно вписать расчетный косинус, если это нужно
        feeder_dict.setdefault('cos', format(feeder_cos, '.2f'))
        # Iрасч - ввод
        # print(sheet, 'Iрасч - ввод')
        try:
            feeder_current = get_float_exel(df, group, 'Iрасч, А')
        except:  # значит будем вычислять
            if U <= 230:
                feeder_current = feeder_power * feeder_use / feeder_cos / U
            else:
                feeder_current = feeder_power * feeder_use / feeder_cos / 1.73205081 / U
                print(feeder_current, '- feeeder_current')
        feeder_dict.setdefault('Iрасч, А',  format(feeder_current, '.1f'))
        # CB - ввод
        # print(sheet, 'CB - ввод')
        try:
            type_cb = get_column_exel(df, group, 'Автомат')[0]  # QF, QFD, QS and add QFS (независ. расцепитель)
        except:
            type_cb = 'QF'
        group_c = ''  # на вводе индекс не делаем
        if U <= 230:  # значение напряжения введено, значит нужно что то поставить
            if type_cb == 'QF':
                pole = 1
                module_counter += 1
            elif type_cb == 'QFD' or type_cb == 'QS':
                pole = 2  # примем что используем двухполюсные ВН
                module_counter += 2
            elif type_cb == 'QFS':
                pole = 1  # автомат остается однополюсным
                module_counter += 2  # но к автомату ставим РН
            else:
                type_cb = 'QF'  # нужно для выбора блока фидера
                pole = 1
                module_counter += 1
        else:
            if type_cb == 'QF':
                pole = 3
                module_counter += 3
            elif type_cb == 'QFD' or type_cb == 'QS':  # примем что используем четырехполюсные ВН
                pole = 4
                module_counter += 4
            elif type_cb == 'QFS':
                pole = 3  # автомат остается трехполюсным
                module_counter += 4  # но к автомату ставим РН
            else:
                type_cb = 'QF'  # нужно для выбора блока фидера
                pole = 3
                module_counter += 3
        feeder_dict.setdefault('type_cb', type_cb)  # нужно для выбора блока фидера
        feeder_dict.setdefault('sf_counter', 0)
        feeder_dict.setdefault('cb', f'{type_cb}{group_c}({pole}p)')
        # I_SB_SET - ввод
        print(sheet, 'I_SB_SET - ввод')
        try:
            i_cb_set = int(get_float_exel(df, group, 'Уставка'))
        except:
            i_cb_max = max(i_cb_max_set)  # находим максимальное значение автомата в данном щите
            c_breakers = [6, 10, 16, 20, 25, 32, 40, 50, 63, 80, 100, 125, 160, 200, 250]  # choosing cb
            good_cb_list = []
            for cb in c_breakers:
                if cb <= i_cb_max:  # найдем автомат больше по значению чем любой отходящий
                    continue
                else:  # затем выберем по току
                    if feeder_current * 1.25 <= cb:
                        good_cb_list.append(cb)
                        break
                    else:
                        continue
            i_cb_set = int(good_cb_list[0])  # значение уставки
        feeder_dict.setdefault('i_cb_set', i_cb_set)  # итог - запись в словарь
        # LOAD_NAME - ввод
        print('LOAD_NAME - ввод')
        load_name = get_column_exel(df, group, 'Нагрузка')[0]
        if type(load_name) == str:  # вводим что написано
            feeder_dict.setdefault('load_name', load_name)
        else:
            feeder_dict.setdefault('load_name', 'Ввод от ВРУ')
        # LEN - ввод
        print(sheet, 'LEN - ввод')
        try:
            length = get_column_exel(df, group, 'Длина, м')[0]
        except:
            length = ''
        try:
            feeder_dict.setdefault('len', int(length))
        except:  # значит не успели посчитать длины кабелей и ввод идет без кабеля
            if type(length) == str:  # написали см. проект ....
                feeder_dict.setdefault('len', length)
            else:
                length = ''
                feeder_dict.setdefault('len', length)
        # CABLE_TYPE - ввод
        if length == '':
            feeder_dict.setdefault('cable', '')
            feeder_dict.setdefault('v_loss', '')
            feeder_dict.setdefault('lay', '')
            feeder_dict.setdefault('lay_l', 0)
            feeder_dict.setdefault('cable', '')
            feeder_dict.setdefault('n_plan', '')
        else:
            try:
                cable_t = get_column_exel(df, group, 'Кабель')[0]
                if type(cable_t) == str:  # это всегда текст. nan - это float, поэтому фильтр не пройдет
                    cable_type = cable_t
                else:
                    cable_type = 'ВВГнг(A)-LS'  # по умолчанию будет данный тип кабеля
            except:
                cable_type = 'ВВГнг(A)-LS'  # по умолчанию будет данный тип кабеля
            # CABLE_SIZE - ввод
            print(sheet, 'CABLE_SIZE - ввод')
            try:
                cable_s = get_float_exel(df, group, 'Сечение')
                if cable_s < 3:  # берем кабели сечением до 2,5мм2, чтобы отобразить их правильно
                    cable_size = format(cable_s, '.1f')
                else:  # берем кабели сечением с 4мм2, чтобы отобразить их правильно
                    cable_size = format(cable_s, '.0f')
                v_loss = 0  # предварительное значение
            except:  # значение не введено в таком случае подбираем кабель
                cable_size_loss = calc_s(U, length, feeder_power, feeder_use, cable_type, i_cb_set, 0)
                cable_size = cable_size_loss[0]  # нашли сечение
                v_loss = cable_size_loss[1]
            if U <= 230:
                cable = f'{cable_type} 3x{cable_size}'
            else:
                cable = f'{cable_type} 5x{cable_size}'
            feeder_dict.setdefault('cable', cable)
            # LAY - ввод
            try:
                type_lay = get_column_exel(df, group, 'Тип прокладки')[0]
            except:
                type_lay = 0  # если столбец удален, то прокладка без трубы
            if type(type_lay) == str:  # если значение введено, то это прокладка в трубе (введено значение П)
                lay = f'{type_lay}{find_lay(cable)}'  # Получается П20 или П25 и тд.
            else:
                lay = ''
            feeder_dict.setdefault('lay', lay)
            # LAY_LENGTH - ввод
            if lay == '':
                feeder_dict.setdefault('lay_l', 0)
            else:
                feeder_dict.setdefault('lay_l', int(length - 1))
            # N_PLAN - ввод
            try:
                n_plan = get_column_exel(df, group, 'N по плану')[0]
            except:
                n_plan = ''
            if type(n_plan) == str:
                feeder_dict.setdefault('n_plan', n_plan)
            else:
                feeder_dict.setdefault('n_plan', '')
            # V_LOSS - ввод
            # print(sheet, 'V_LOSS - ввод')
            try:
                feeder_loss = get_float_exel(df, group, 'dU')
                feeder_dict.setdefault('v_loss', format(feeder_loss, '.2f'))
            except:
                if v_loss > 0:
                    feeder_dict.setdefault('v_loss', format(v_loss, '.2f'))
                else:
                    v_loss = calc_dv(  # считаем потери
                        U, length, feeder_power, feeder_use, cable_type, float(cable_size))
                    feeder_dict.setdefault('v_loss', format(v_loss, '.2f'))
        feeders_list.insert(0, feeder_dict)  # добавляем инфо о вводе в список с фидерами на 0 место
        panel_list.append(feeders_list)  # add info about feeder to panel list (number 0)
        panel_list.append(quantity)  # add quantity to panel list (number 1)
        panel_list.append(sheets[sheet])  # add panel name (number 2)
        panel_counter += 1
        if module_counter < 8:
            panel_box = 'ЩРн-12з-0'
        elif module_counter < 14:
            panel_box = 'ЩРн-18з-0'
        elif module_counter < 20:
            panel_box = 'ЩРн-24з-0'
        elif module_counter < 32:
            panel_box = 'ЩРн-36з-0'
        elif module_counter < 44:
            panel_box = 'ЩРн-48з-0'
        elif module_counter < 50:
            panel_box = 'ЩРн-54з-0'
        else:
            panel_box = 'ЩРн-72з-0'
        panel_list.append(panel_box)  # number 3
        panel_list.append(pnom)  # number 4
        power_calc = format(sum_power * feeder_use / 1000,
                            '.2f')  # feeder_use - здесь коэфициент использования ввода
        panel_list.append(power_calc)  # add расчетную мощность (number 5)
        panel_list.append(format(feeder_current, '.1f'))  # add расчетный ток щита см. Iрасч - ввод, (number 6)
        panel_list.append(format(feeder_cos, '.2f'))  # коэф мощности щита (number 7)
        panel_list.append(format(feeder_use, '.2f'))  # (number 8)
        panel_list.append(place_counter)  # (number 9) - количество места на листе (нужно для выбора листа А3 или А2)
        print(quantity, place_counter)
        panels_list.append(panel_list)  # add info about panel to panels list
        print(panel_list)

    return panels_list


def lay_count(panels_list):
    lays_data = []  # список со словарями труб для каждой панели
    for panel_list in panels_list:
        lay_dict = {  # трубы, используемые в панели (можно добавить)
            'П16': 0,
            'П20': 0,
            'П25': 0,
            'П32': 0,
            'П40': 0,
            'П50': 0,
            'П63': 0,
        }
        feeders_list = panel_list[0]  # get list of dicts with feeder info
        for feeder in feeders_list[1:]:
            print(feeder)
            feeder_lay = feeder.get('lay')  # взяли lay фидера
            lay_length = lay_dict.get(feeder_lay)  # взяли длину этого lay для этой панели
            if not feeder_lay == None or not lay_length == None:  # проверка на наличие lay и корректности lay
                feeder_lay_len = feeder.get('lay_l')  # взяли длину этого lay для этого фидера
                if type(feeder_lay_len) == int:  # все значения длин типа int
                    new_length = lay_length + feeder_lay_len
                    lay_dict[feeder_lay] = new_length
                    sf_counter = feeder.get('sf_counter')
                    if sf_counter > 0:  # если есть подфидеры
                        for sf in range(1, sf_counter + 1):  # если есть подфидеры
                            subfeeder = feeder.get(f'subfeeder_{sf}')  # взяли словарь подфидера
                            subfeeder_lay = subfeeder.get('lay')  # взяли lay подфидера
                            lay_length = lay_dict.get(subfeeder_lay)  # взяли длину этого lay для этой панели
                            print(lay_length)
                            if not subfeeder_lay == None and not lay_length == None:  # проверка на наличие lay и lay
                                subfeeder_lay_len = subfeeder.get('lay_l')
                                if type(subfeeder_lay_len) == int:
                                    lay_dict.setdefault(subfeeder_lay, lay_length + subfeeder_lay_len)
                                else:
                                    continue
                            else:
                                print('hi')
                                continue
                    else:
                        continue
                else:
                    continue
            else:
                continue

        lays_data.append(lay_dict)
    # итоговый словарь с трубами
    lays_dict = {  # трубы, используемые в проекте (можно добавить)
        'П16': 0,
        'П20': 0,
        'П25': 0,
        'П32': 0,
        'П40': 0,
        'П50': 0,
        'П63': 0,
    }
    for lay_panel in lays_data:  # берем cловать с прокладкой для каждой панели
        lays_dict = Counter(lays_dict) + Counter(lay_panel)

    return lays_data, lays_dict


def fill_lay_data(atts, lay_data, panel_counter):
    atts_list = list(atts)
    print('atts_list', atts_list)
    lay_dict = lay_data[panel_counter]  # получили словарь с трубами для этой панели
    lay_types = ['П16', 'П20', 'П25', 'П32', 'П40', 'П50', 'П63']  # трубы, используемые в проекте (можно добавить)
    c = 1
    for lay_type in lay_types:
        len_lay = lay_dict[lay_type]
        print('len_lay', len_lay, type(len_lay))
        if type(len_lay) == int and len_lay > 0:
            for att in atts_list:  # находим атрибут, обозначающий название столбца в таблице кабелей
                if att.TagString == f'TUBE{c}':
                    att.TextString = lay_type
                    break
                else:
                    continue
            for att in atts_list:  # находим атрибут, обозначающий название столбца в таблице кабелей
                if att.TagString == f'D{c}':
                    l_type = lay_type.split('П')[1]
                    att.TextString = l_type
                    break
                else:
                    continue
            for att in atts_list:
                if att.TagString == f'L{c}':
                    att.TextString = len_lay
                else:
                    continue
            c += 1
        else:
            continue

def fill_cable_data(atts, cable_data, panel_counter):
    atts_list = list(atts)  # list of attributes of the block
    cable_list = cable_data[panel_counter]  # взяли список с инфо о кабелях нужной панели
    # доп скрипт для авто наименования столбцов с типом кабеля
    cable_type_list = [
        'ВВГнг(A)-LS', 'ВВГнг(A)-FRLS', 'АВВГнг(A)-LS', 'АсВВГнг(A)-LS', 'ВВГнг(A)-LSLTx', 'ВВГнг(A)-FRLSLTx',
        'ВВГнг(A)-HF', 'ВВГнг(A)-FRHF'
    ]
    c = 0  # счетчик столбца (1 или 2)
    n = 0  # счетчик типа кабеля
    for cab in cable_list:  # берем каждый элемент списка (это словарь или 0) с типами кабелями
        if c == 2:  # в щите может быть только 2 типа кабеля, так как в таблице всего 2 столбца
            break
        else:
            if not cab == 0:  # если это словарь, то
                c += 1  # поднимаем номер столбца сразу как находим наличие кабеля
                for att in atts_list:  # находим атрибут, обозначающий название столбца в таблице кабелей
                    if att.TagString == f'TYPE{c}':
                        typ_c = cable_type_list[n]
                        att.TextString = typ_c  # вводим значение данного сечения
                        break
                    else:
                        continue
                n += 1  # заканчиваем работу с данным типом - переходим к следующему
            else:
                n += 1  # заканчиваем работу с данным типом - переходим к следующему
    zero_count = cable_list.count(0)  # считаем сколько 0 в списке
    for i in range(0, zero_count):
        print(i)
        cable_list.remove(0)  # удаляем все 0 в списке с кабелями
    if len(cable_list) < 2:
        cable_list.append(0)
        cable_list.append(0)
    # конец работы доп скрипта
    cables_list = ['3x1.5', '3x2.5', '3x4', '3x6', '3x10', '3x16', '5x1.5', '5x2.5', '5x4', '5x6', '5x10', '5x16']
    types_list = ['VVGLS', 'VVGFRLS']  # на самом деле это [1TYPE, 2TYPE]
    str_counter = 1  # счетчик строки в таблице, куда вносятся значения
    for key in cables_list:  # берем кабель
        size_counter = 0  # cчитает сколько кабелей данного сечения удалось зафиксировать в разных типах кабеля
        for i in range(0, 2):  # будем брать каждый словарь и проверять наличие данного кабеля в нем
            cable_dict = cable_list[i]  # взяли dict для опреденного типа кабеля (или 0 если типа в щите нет)
            if not cable_dict == 0:  # значит данный тип кабеля есть в данном щите и нужно смотреть сечения
                if cable_dict[key] > 0:  # кабель данного сечения и данного типа есть
                    size_counter += 1  # пишем что кабель зафиксировали
                    for att in atts_list:  # находим атрибут, обозначающий название строки в таблице кабелей
                        if att.TagString == f'CABLE{str_counter}':
                            att.TextString = key  # вводим значение данного сечения
                        else:
                            continue
                    att_name = types_list[i]  # нужен атрибут кабеля
                    for att in atts_list:  # находим второй атрибут для кабеля и меняем его
                        if att.TagString == f'{att_name}{str_counter}':
                            att.TextString = cable_dict.get(key)
                        else:
                            continue
                else:
                    continue
            else:  # если данного типа кабеля нет, то проверяем это сечение в других типах
                continue
        if size_counter > 0:  # если хотя бы один кабель данного сечения удалось обнаружить, то меняем строку таблицы
            str_counter += 1

def bom_cable_count(cable_data):  # input [[{}, {}, 0, 0, 0, 0, 0], [{}, {}, 0, 0, 0, 0, 0],....]
    bom_cable_data = [0, 0, 0, 0, 0, 0, 0, 0]
    cable_dict = {  # начальный словарь с нулевыми длинами кабеля
        '3x1.5': 0, '3x2.5': 0, '3x4': 0, '3x6': 0, '3x10': 0, '3x16': 0, '3x25': 0, '3x35': 0,
        '3x50': 0,
        '5x1.5': 0, '5x2.5': 0, '5x4': 0, '5x6': 0, '5x10': 0, '5x16': 0, '5x25': 0, '5x35': 0,
        '5x50': 0,
        '5x70': 0, '5x95': 0, '5x120': 0, '5x150': 0, '5x185': 0, '5x240': 0
    }  # типы кабеля указаны сразу, чтобы соблюсти порядок
    for i in range(0, 8):  # у нас 8 типов кабеля, берем один за другим
        for types_list in cable_data:  # берем лист для каждой панели, который состоит из 0 или списков
            type_dict = types_list[i]  # берем списки (или 0) одного типа кабеля
            if type_dict == 0:  # значит такого типа в панели нет
                continue
            else:
                bom_c_d = bom_cable_data[i]
                if bom_c_d == 0:
                    dict_1 = Counter(cable_dict)
                    dict_2 = Counter(type_dict)
                    new_cable_dict = dict_1 + dict_2
                    bom_cable_data[i] = new_cable_dict  # обновляем словарь
                else:
                    dict_1 = Counter(bom_c_d)
                    dict_2 = Counter(type_dict)
                    new_cable_dict = dict_1 + dict_2
                    bom_cable_data[i] = new_cable_dict  # обновляем словарь
    return bom_cable_data


# BOM_MODULE
def cable_writer(cable_list, i, cable_size, feeder):
    cable_dict = {  # начальный словарь с нулевыми длинами кабеля
        '3x1.5': 0, '3x2.5': 0, '3x4': 0, '3x6': 0, '3x10': 0, '3x16': 0, '3x25': 0, '3x35': 0,
        '3x50': 0,
        '5x1.5': 0, '5x2.5': 0, '5x4': 0, '5x6': 0, '5x10': 0, '5x16': 0, '5x25': 0, '5x35': 0,
        '5x50': 0,
        '5x70': 0, '5x95': 0, '5x120': 0, '5x150': 0, '5x185': 0, '5x240': 0
    }  # типы кабеля указаны сразу, чтобы соблюсти порядок
    if not cable_list[i] == 0:  # если такой кабель уже был в щите, то берем словарь из списка с кабелями
        cable_type_dict = cable_list[i]
    else:
        cable_list[i] = cable_dict  # заменяем 0 на dict, что значит - такой тип кабель в щите есть
        cable_type_dict = cable_list[i]
    if cable_size in cable_type_dict:
        old_length = cable_type_dict.get(cable_size)  # take vvg_ls cable type
        new_length = old_length + int(feeder.get('len'))
        cable_type_dict[cable_size] = new_length
    return cable_type_dict  # возвращает словарь с актульаными длинами кабеля

def cable_count(panels_list):  # возращает list из listов вида [[{}, {}, 0, 0, 0, 0, 0, 0], [{}, {}, 0, 0, 0, 0, 0, 0]]
    cable_data = []
    for panel_list in panels_list:  # take every panel
        cable_list = [  # лист со словарями или нулями, индекс каждого принадлежит определ-му типу кабеля
            0, 0, 0, 0, 0, 0, 0, 0
        ]  # index = 0-ВВГнгLS, 1-ВВГнгFRLS, 2-АВВГнг-LS, 3-АВВГнг(A)-LS, 4-ВВГнг-LSLTx,
        # 5-ВВГнг(A)-FRLSLTx index = 6-ВВГнг(A)-HF, 7-ВВГнг(A)-FRHF
        feeders_list = panel_list[0]  # get list of dicts with feeder info
        cable_type_list = [
            'ВВГнг(A)-LS', 'ВВГнг(A)-FRLS', 'АВВГнг(A)-LS', 'АсВВГнг(A)-LS', 'ВВГнг(A)-LSLTx', 'ВВГнг(A)-FRLSLTx',
            'ВВГнг(A) - HF', 'ВВГнг(A)-FRHF'
        ]
        for feeder in feeders_list[1:]:  # take every feeder except Input_CB
            print(feeder)
            feeder_cable = feeder.get('cable')  # get feeder's cable
            feeder_cable_list = feeder_cable.split()  # get list with cable type and length
            if len(feeder_cable_list) > 1:
                cable_type = feeder_cable_list[0]
                cable_size = feeder_cable_list[1]
                i = 0  # индекс типа кабеля, который рассматривается в данном фидере
                for cable_t in cable_type_list:  # идет распределение по типам кабеля
                    if cable_type == cable_t:
                        cable_list_new = cable_writer(cable_list, i, cable_size, feeder)
                        cable_list[i] = cable_list_new  # обновили словарь
                        sf_counter = feeder.get('sf_counter')
                        if sf_counter > 0:  # если у фидера есть подфидеры
                            for sf in range(1, sf_counter + 1):
                                subfeeder = feeder.get(f'subfeeder_{sf}')  # взяли словарь с подфидером
                                print(subfeeder)
                                subfeeder_cable = subfeeder.get('cable')  # взяли кабель подфидера
                                print(subfeeder_cable)
                                subfeeder_cable_list = subfeeder_cable.split()  # разделили на тип и на сечение
                                if len(subfeeder_cable_list) > 1:
                                    cable_size = subfeeder_cable_list[1]
                                    cable_list_new = cable_writer(cable_list, i, cable_size, subfeeder)
                                    cable_list[i] = cable_list_new  # обновили словарь
                                else:
                                    continue
                        break
                    else:
                        i += 1  # переход к следующей итерации и изменение индекса типа кабеля
            else:  # кабель может быть комплектным или не указан
                continue
        cable_data.append(cable_list)

    return cable_data

# CALC_MODULE
def calc_dv(U, length, feeder_power, feeder_use, cable_type, s_cable):
    moment = length * feeder_power / 1000 * feeder_use
    if cable_type == 'ВВГнг(A)-LS' or cable_type == 'ВВГнг(A)-FRLS':
        if U <= 230:
            d_v = moment / 12.1 / s_cable
        else:
            d_v = moment / 72.4 / s_cable
    else:
        if U <= 230:
            d_v = moment / 7.4 / s_cable
        else:
            d_v = moment / 44 / s_cable
    return d_v

def calc_cb(feeder_current):
    c_breakers = [6, 10, 16, 20, 25, 32, 40, 50, 63, 80, 100, 125, 160, 200, 250]  # choosing cb
    good_cb_list = []
    for cb in c_breakers:
        if feeder_current * 1.25 <= cb:
            good_cb_list.append(cb)
            break
        else:
            continue
    return int(good_cb_list[0])

def calc_s(U, length, feeder_power, feeder_use, cable_type, i_cb_set, dv_max):  # возращает лист с сечением кабеля
    if cable_type == 'АВВГнг(A)-LS' or cable_type == 'АсВВГнг(A)-LS':
        if U <= 230:
            cables = {  # the dict for relation CB and current
                6: [1.5, 2.5, 4, 6, 10, 16, 25, 35, 50, 70, 95, 120, 150, 185, 240],
                10: [2.5, 4, 6, 10, 16, 25, 35, 50, 70, 95, 120, 150, 185, 240],
                16: [2.5, 4, 6, 10, 16, 25, 35, 50, 70, 95, 120, 150, 185, 240],
                20: [4, 6, 10, 16, 25, 35, 50, 70, 95, 120, 150, 185, 240],
                25: [6, 10, 16, 25, 35, 50, 70, 95, 120, 150, 185, 240],
                32: [6, 10, 16, 25, 35, 50, 70, 95, 120, 150, 185, 240],
                40: [10, 16, 25, 35, 50, 70, 95, 120, 150, 185, 240],
                50: [16, 25, 35, 50, 70, 95, 120, 150, 185, 240],
                63: [25, 35, 50, 70, 95, 120, 150, 185, 240],
                80: [35, 50, 70, 95, 120, 150, 185, 240],
                100: [50, 70, 95, 120, 150, 185, 240],
                125: [50, 70, 95, 120, 150, 185, 240],
                160: [95, 120, 150, 185, 240],
                200: [120, 150, 185, 240],
            }
        else:
            cables = {  # the dict for relation CB and current
                6: [1.5, 2.5, 4, 6, 10, 16, 25, 35, 50, 70, 95, 120, 150, 185, 240],
                10: [2.5, 4, 6, 10, 16, 25, 35, 50, 70, 95, 120, 150, 185, 240],
                16: [4, 6, 10, 16, 25, 35, 50, 70, 95, 120, 150, 185, 240],
                20: [6, 10, 16, 25, 35, 50, 70, 95, 120, 150, 185, 240],
                25: [6, 10, 16, 25, 35, 50, 70, 95, 120, 150, 185, 240],
                32: [10, 16, 25, 35, 50, 70, 95, 120, 150, 185, 240],
                40: [16, 25, 35, 50, 70, 95, 120, 150, 185, 240],
                50: [25, 35, 50, 70, 95, 120, 150, 185, 240],
                63: [25, 35, 50, 70, 95, 120, 150, 185, 240],
                80: [50, 70, 95, 120, 150, 185, 240],
                100: [50, 70, 95, 120, 150, 185, 240],
                125: [70, 95, 120, 150, 185, 240],
                160: [95, 120, 150, 185, 240],
                200: [150, 185, 240],
            }
    else:
        if U <= 230:
            cables = {  # the dict for relation CB and current
                6: [1.5, 2.5, 4, 6, 10, 16, 25, 35, 50, 70, 95, 120, 150, 185, 240],
                10: [1.5, 2.5, 4, 6, 10, 16, 25, 35, 50, 70, 95, 120, 150, 185, 240],
                16: [2.5, 4, 6, 10, 16, 25, 35, 50, 70, 95, 120, 150, 185, 240],
                20: [2.5, 4, 6, 10, 16, 25, 35, 50, 70, 95, 120, 150, 185, 240],
                25: [4, 6, 10, 16, 25, 35, 50, 70, 95, 120, 150, 185, 240],
                32: [6, 10, 16, 25, 35, 50, 70, 95, 120, 150, 185, 240],
                40: [10, 16, 25, 35, 50, 70, 95, 120, 150, 185, 240],
                50: [10, 16, 25, 35, 50, 70, 95, 120, 150, 185, 240],
                63: [16, 25, 35, 50, 70, 95, 120, 150, 185, 240],
                80: [25, 35, 50, 70, 95, 120, 150, 185, 240],
                100: [25, 35, 50, 70, 95, 120, 150, 185, 240],
                125: [35, 50, 70, 95, 120, 150, 185, 240],
                160: [50, 70, 95, 120, 150, 185, 240],
                200: [95, 120, 150, 185, 240],
            }
        else:
            cables = {  # the dict for relation CB and current
                6: [1.5, 2.5, 4, 6, 10, 16, 25, 35, 50, 70, 95, 120, 150, 185, 240],
                10: [1.5, 2.5, 4, 6, 10, 16, 25, 35, 50, 70, 95, 120, 150, 185, 240],
                16: [2.5, 4, 6, 10, 16, 25, 35, 50, 70, 95, 120, 150, 185, 240],
                20: [2.5, 4, 6, 10, 16, 25, 35, 50, 70, 95, 120, 150, 185, 240],
                25: [4, 6, 10, 16, 25, 35, 50, 70, 95, 120, 150, 185, 240],
                32: [6, 10, 16, 25, 35, 50, 70, 95, 120, 150, 185, 240],
                40: [10, 16, 25, 35, 50, 70, 95, 120, 150, 185, 240],
                50: [16, 25, 35, 50, 70, 95, 120, 150, 185, 240],
                63: [16, 25, 35, 50, 70, 95, 120, 150, 185, 240],
                80: [25, 35, 50, 70, 95, 120, 150, 185, 240],
                100: [35, 50, 70, 95, 120, 150, 185, 240],
                125: [50, 70, 95, 120, 150, 185, 240],
                160: [70, 95, 120, 150, 185, 240],
                200: [120, 150, 185, 240],
            }
    good_s_cable_list = []  # отбираются значения, проходящие по току
    a = cables.get(i_cb_set)
    for s_cable in a:  # take first cable
        d_v = calc_dv(U, length, feeder_power, feeder_use, cable_type, s_cable) + dv_max
        if d_v <= 3.0:
            good_s_cable_list.append(s_cable)  # сечение кабеля
            good_s_cable_list.append(d_v)
            break
        else:
            continue
    return good_s_cable_list  # возращает лист со значением сечения кабеля и потерями в кабеле

# READER_MODULE
def get_column_exel(df, group, column):
    return df.loc[df['Группа'] == group, column].tolist()


def get_float_exel(df, group, column):
    value = df.loc[df['Группа'] == group, column].tolist()[0]
    if float(value) > 0:   # проверка на число, и то что оно больше нуля
        int(value)
        return float(value)
    else:  # это вызовет ошибку если если число = 0
        var = 1 / int(value)
        return var

def find_lay(cable):
    cable_d = cable.split()[1]
    tubes = {
        '16': ['3x1.5'],
        '20': ['5x1.5', '3x2.5', '3x4'],
        '25': ['5x2.5', '5x4', '3x6'],
        '32': ['5x6', '3x10', '3x16', '5x10'],
        '40': ['5x16', '3x25'],
        '50': ['5x25', '3x35', '3x50'],
        '63': ['3x70'],
    }

    for key in tubes:  # берем все ключи словаря
        s = tubes[key]  # получем лист по ключу
        if s.count(cable_d) > 0:  # если в списке есть
            return key
        else:
            continue

sb_production = st.radio(
    'Выберите производителя оборудования',
    ['IEK', 'EKF']
)

file_exel = st.file_uploader("Choose a file")


if file_exel is not None:
    name_file = file_exel.name.split('.')[0]
    # name_file = file_exel.split('/')[-1]  # link
    panels_list = get_info_exel(file_exel)  # get list of lists with panel info
    print('exit from reader')
    cable_data = cable_count(panels_list)
    print('exit from cable_data')
    lay_data = lay_count(panels_list)
    print('exit from lay_data')
    print('Lay_data', lay_data)
    # запись данных в txt
    bom_cable = bom_cable_count(cable_data)
    print("список кабелей идущих в TXT-", bom_cable)
    with open('C:\\AutoSLD\\bom.txt', 'a') as f:
        f.write(f'\n\nИТОГОВАЯ ДЛИНА КАБЕЛЯ В ФАЙЛЕ {name_file}')
        # f.write(f'\n\nИТОГОВАЯ ДЛИНА КАБЕЛЯ В ФАЙЛЕ STREAMLIT')
        cable_type_list = [
            'ВВГнг(A)-LS', 'ВВГнг(A)-FRLS', 'АВВГнг(A)-LS', 'АсВВГнг(A)-LS', 'ВВГнг(A)-LSLTx',
            'ВВГнг(A)-FRLSLTx',  'ВВГнг(A) - HF', 'ВВГнг(A)-FRHF'
        ]
        i = 0
        for cable_type in bom_cable:
            if cable_type == 0:
                i += 1
                continue
            else:
                f.write(f'\n'
                        f'{cable_type_list[i]} - {cable_type}'
                        )
                i += 1
        # f.write(f'\nПОТРЕБНОСТЬ ТРУБ в файле {name_file}-{lay_data[1]}')
        f.write(f'\nПОТРЕБНОСТЬ ТРУБ в файле STREAMLIT')
    print('AutoCAD() start')
    pythoncom.CoInitializeEx(0)
    acad = AutoCAD()
    print('AutoCAD() finish')
    open_file = acad.Documents.Add('test')  # create a new file
    print('new drawing')
    mp = open_file.ModelSpace
    a2_counter = 0  # counter of panel which is on A2
    a3_counter = 0  # counter of panel which is on A3
    for panel_list in panels_list:
        feeders_list = panel_list[0]  # list of dicts with feeder info
        quantity = panel_list[1]  # quantity of feeders
        panel_name = panel_list[2]
        box_name = panel_list[3]
        sum_power = panel_list[4]
        power_calc = panel_list[5]
        calc_current = panel_list[6]
        panel_cos = panel_list[7]
        panel_use = panel_list[8]
        panel_place = panel_list[9]

        # choose and insert a frame
        if panel_place > 24:
            point_frame = AcadPoint(100000, int(0 - a2_counter * 60000), 0)  # point of initial insert
            mp.InsertBlock(
                point_frame(),
                'C:\\AutoSLD\\frames\\title_block_A2.dwg',
                1.0, 1.0, 1.0, 0
            )  # insert the frame
            point_title = AcadPoint(141500, int(-58900 - a2_counter * 60000), 0)
            title_blk = mp.InsertBlock(point_title(),
                                       'C:\\AutoSLD\\title_blocks\\aqmol_new.dwg',
                                       1.0, 1.0, 1.0, 0)  # insert the title
            atts = title_blk.GetAttributes()  # array of objects Attribute
            # fill_title(atts, title)  # fill title block
            page_name = f'Однолинейная схема {panel_name}'
            n_list = f'{a2_counter + a3_counter + 1}'
            for att in list(atts):
                att_name = att.TagString
                if att_name == 'N_LIST':
                    att.TextString = n_list
                elif att_name == 'PAGE_NAME':
                    att.TextString = page_name  # f'Однолинейная схема {panel_name}'
            # вставляем табличку потребность труб
            point_lay_data = AcadPoint(110700, int(-58900 - a2_counter * 60000), 0)
            lay_data_blk = mp.InsertBlock(point_lay_data(),
                                            'C:\\AutoSLD\\bom_data\\lay_data_A2.dwg',
                                            1.0, 1.0, 1.0, 0)  # insert the table with cable data
            atts = lay_data_blk.GetAttributes()
            fill_lay_data(atts, lay_data[0], a2_counter + a3_counter)
            # вставляем табличку с подсчетом кабеля
            point_cable_data = AcadPoint(102000, int(-58900 - a2_counter * 60000), 0)
            cable_data_blk = mp.InsertBlock(point_cable_data(),
                                            'C:\\AutoSLD\\bom_data\\cable_data_A2.dwg',
                                            1.0, 1.0, 1.0, 0)  # insert the table with cable data
            atts = cable_data_blk.GetAttributes()
            fill_cable_data(atts, cable_data, a2_counter + a3_counter)
            # вставляем информацию о нагрузке панели
            point_panel_data = AcadPoint(105000, int(-11700 - a2_counter * 60000), 0)
            panel_info_blk = mp.InsertBlock(point_panel_data(),
                                            'C:\\AutoSLD\\blocks\\panel_info.dwg',
                                            1.0, 1.0, 1.0, 0)  # insert the table with cable data
            a2_counter += 1
            x_coord = 100000
        else:
            point_frame = AcadPoint(0, int(0 - a3_counter * 30000), 0)  # point of initial insert
            mp.InsertBlock(
                point_frame(),
                'C:\\AutoSLD\\frames\\title_block_A3.dwg',
                1.0, 1.0, 1.0, 0)  # insert the block
            point_title = AcadPoint(41500, int(-29200 - a3_counter * 30000), 0)
            title_blk = mp.InsertBlock(point_title(),
                                       'C:\\AutoSLD\\title_blocks\\aqmol_new.dwg',
                                       1.0, 1.0, 1.0, 0)  # insert the title
            atts = title_blk.GetAttributes()  # array of objects Attribute
            # fill_title(atts, title)  # fill title block
            page_name = f'Однолинейная схема {panel_name}'
            n_list = f'{a2_counter + a3_counter + 1}'
            for att in list(atts):
                # print(att.TagString)
                if att.TagString == 'N_LIST':
                    att.TextString = n_list
                elif att.TagString == 'PAGE_NAME':
                    att.TextString = page_name  # f'Однолинейная схема {panel_name}'
            # вставляем табличку потребность труб
            point_lay_data = AcadPoint(10700, int(-29200 - a3_counter * 30000), 0)
            lay_data_blk = mp.InsertBlock(point_lay_data(),
                                          'C:\\AutoSLD\\bom_data\\lay_data_A3.dwg',
                                          1.0, 1.0, 1.0, 0)  # insert the table with cable data
            atts = lay_data_blk.GetAttributes()
            fill_lay_data(atts, lay_data[0], a2_counter + a3_counter)
            # вставляем табличку с кабелями
            point_cable_data = AcadPoint(2000, int(-29200 - a3_counter * 30000), 0)
            cable_data_blk = mp.InsertBlock(
                point_cable_data(),
                'C:\\AutoSLD\\bom_data\\cable_data_A3.dwg',
                1.0, 1.0, 1.0, 0)  # insert the table with cable data
            atts = cable_data_blk.GetAttributes()
            fill_cable_data(atts, cable_data, a2_counter + a3_counter)
            point_panel_data = AcadPoint(5000, int(-11700 - a3_counter * 30000), 0)
            panel_info_blk = mp.InsertBlock(
                point_panel_data(),
                'C:\\AutoSLD\\blocks\\panel_info.dwg',
                1.0, 1.0, 1.0, 0)  # insert the table with cable data
            a3_counter += 1
            x_coord = 0
        atts = panel_info_blk.GetAttributes()
        for att in list(atts):
            if att.TagString == 'PANEL_NAME':
                att.TextString = panel_name
            elif att.TagString == 'BOX_NAME':
                att.TextString = box_name
            elif att.TagString == 'SUM_POWER':
                att.TextString = f'Pу={sum_power} кВт'
            elif att.TagString == 'POWER_CALC':
                att.TextString = f'Pр={power_calc} кВт'
            elif att.TagString == 'CALC_CURRENT':
                att.TextString = f'Iр={calc_current} A'
            elif att.TagString == 'PANEL_COS':
                att.TextString = f'Cosф={panel_cos}'
            elif att.TagString == 'PANEL_USE':
                att.TextString = f'Kc={panel_use}'
            else:
                continue
        if x_coord == 100000:
            point0 = AcadPoint(108868.5248, int(-5603.0806 - (a2_counter - 1) * 60000), 0)  # point of first feeder
        else:
            point0 = AcadPoint(8868.5248, int(-5603.0806 - (a3_counter - 1) * 30000), 0)  # point of first feeder
        subfeeder_counter = 1
        for f in range(1, quantity + 1):
            consider_feeder = feeders_list[f - 1]
            # print(consider_feeder)
            type_cb = consider_feeder.get('type_cb')  # у каждого фидера должен быть указан тип
            sf_counter = consider_feeder.get('sf_counter')  # и у каждого кол-во подфидеров
            nextpoint = (0, 700, 0)  # start adding feeders
            delta_point = AcadPoint(nextpoint)  # difference between the feeder and point of initial insert
            point = point0 - subfeeder_counter * delta_point  # point between feeders
            blk = mp.InsertBlock(
                point(), f'C:\\AutoSLD\\feeders\\{type_cb}_{sf_counter}sf.dwg',
                1.0, 1.0, 1.0, 0
            )  # insert the block
            if sf_counter == 0:
                subfeeder_counter += 1
            else:
                subfeeder_counter += sf_counter
            atts = blk.GetAttributes()  # array of objects Attribute
            consider_subfeeder_list = []
            for i in range(1, sf_counter + 1):
                consider_subfeeder_list.append(consider_feeder.get(f'subfeeder_{i}'))
            for att in list(atts):
                tag = att.TagString
                # print(panel_name, f'feeder{f}', tag)
                if tag == 'FEEDER':
                    att.TextString = consider_feeder.get('feeder')
                elif tag == 'CB':
                    att.TextString = consider_feeder.get('cb')
                elif tag == 'I_CB_SET':
                    att.TextString = consider_feeder.get('i_cb_set')
                elif tag == 'CABLE':
                    att.TextString = consider_feeder.get('cable')
                elif tag == 'LEN':
                    att.TextString = consider_feeder.get('len')
                elif tag == 'LAY':
                    att.TextString = consider_feeder.get('lay')
                elif tag == 'N_PLAN':
                    att.TextString = consider_feeder.get('n_plan')
                elif tag == 'PNOM':
                    att.TextString = consider_feeder.get('pnom')
                elif tag == 'INOM':
                    att.TextString = consider_feeder.get('Iрасч, А')
                elif tag == 'V_LOSS':
                    att.TextString = consider_feeder.get('v_loss')
                elif tag == 'LOAD_NAME':
                    att.TextString = consider_feeder.get('load_name')
                elif tag == 'I_SB_NOM':
                    continue
                else:
                    for i in range(1, sf_counter + 1):
                        consider_subf = consider_subfeeder_list[i - 1]
                        if tag == f'CABLE_{i}':
                            att.TextString = consider_subf.get('cable')
                        elif tag == f'LEN_{i}':
                            att.TextString = consider_subf.get('len')
                        elif tag == f'LAY_{i}':
                            att.TextString = consider_subf.get('lay')
                        elif tag == f'N_PLAN_{i}':
                            att.TextString = consider_subf.get('n_plan')
                        elif tag == f'PNOM_{i}':
                            att.TextString = consider_subf.get('pnom')
                        elif tag == f'INOM_{i}':
                            att.TextString = consider_subf.get('Iрасч, А')
                        elif tag == f'V_LOSS_{i}':
                            att.TextString = consider_subf.get('v_loss')
                        elif tag == f'LOAD_NAME_{i}':
                            att.TextString = consider_subf.get('load_name')
                        else:
                            continue
    print('close')
    # print(file_exel.id)
    acad.ActiveDocument.Close('TRUE', f'C:\\Users\\s.savelyev\\Documents\\{name_file}.dwg')
    # data = requests.get('C:\\Users\\s.savelyev\\Documents\\park.dwg')
    with open(f'C:\\Users\\s.savelyev\\Documents\\{name_file}.dwg', 'rb') as dwgf:
        btn = st.download_button(
        label = "Download",
        data = dwgf,
        file_name = f"{name_file}.dwg",
        mime = "dwg"
        )
    os.remove(f'C:\\Users\\s.savelyev\\Documents\\{name_file}.dwg')


    # st.download_button('Download', b'C:\\Users\\s.savelyev\\Documents\\park.dwg')


    # st.download_button('Download', 'C:\\Users\\s.savelyev\\Documents\\park.dwg')


    st.success('Success message')

else:
    st.success('no file')


# st.stop()

